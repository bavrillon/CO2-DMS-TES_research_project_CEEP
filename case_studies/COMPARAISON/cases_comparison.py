"""
This module provides functions to create a synthesis Excel file 
from case study results, and to plot comparisons of variables across 
different cases and TES types.
"""
import pandas as pd
import numpy as np
import os
import sys
import re
import shutil
import matplotlib.pyplot as plt

MAIN_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.append(MAIN_DIR)
from CONSTANTS import *

SYNTHESIS_MODEL_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "case_studies_synthesis_0.xlsx")
SYNTHESIS_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "case_studies_synthesis.xlsx")
FIXED_SYNTHESIS_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "case_studies_synthesis_fixed.xlsx")
PLOTS_SAVING_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cases_comparison_plots")

def compute_case_studies_synthesis(input_dir=CASE_STUDIES_SAVING_DIR, saving_path=SYNTHESIS_FILE_PATH):
    """
    Create a synthesis Excel file from the model template and fill the
    "OPTIMIZATION RESULTS" section with data from case studies.

    Parameters
    ----------
    input_dir : str
        Directory containing case study subfolders.
    saving_path : str
        Path where the synthesis file is saved.

    Returns
    -------
    str
        The path to the generated synthesis file.
    """
    if not os.path.exists(SYNTHESIS_MODEL_FILE_PATH):
        raise FileNotFoundError(f"Template not found: {SYNTHESIS_MODEL_FILE_PATH}")

    os.makedirs(os.path.dirname(saving_path) or ".", exist_ok=True)
    shutil.copyfile(SYNTHESIS_MODEL_FILE_PATH, saving_path)

    this_folder_name = os.path.basename(os.path.abspath(os.path.dirname(__file__)))
    cases = [
        item for item in os.listdir(input_dir)
        if os.path.isdir(os.path.join(input_dir, item)) and item != this_folder_name
    ]
    cases_sorted = sorted(cases)
    if "reference" in cases_sorted:
        cases_sorted = ["reference"] + [case for case in cases_sorted if case != "reference"]

    cases_data = {}
    for case in cases_sorted:
        excel_path = os.path.join(input_dir, case, "outputs_optim.xlsx")
        if not os.path.exists(excel_path):
            cases_data[case] = {}
            continue

        try:
            scalars_df = pd.read_excel(excel_path, sheet_name="Scalars")
            if "variable" in scalars_df.columns and "value" in scalars_df.columns:
                cases_data[case] = scalars_df.set_index("variable")["value"].to_dict()
            else:
                cases_data[case] = {}
        except Exception:
            cases_data[case] = {}

    from openpyxl import load_workbook

    wb = load_workbook(saving_path)
    marker = None
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().casefold() == "optimization results":
                    marker = (ws, cell.row, cell.column)
                    break
            if marker:
                break
        if marker:
            break

    if marker is None:
        raise ValueError("'OPTIMIZATION RESULTS' section not found in synthesis template.")

    ws, marker_row, marker_col = marker
    header_row = None
    variable_col = None
    search_end = marker_row + 10
    keywords = {"variable", "variables", "parametre", "parametres", "paramètre", "paramètres"}

    for r in range(marker_row + 1, search_end + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().casefold() in keywords:
                header_row = r
                variable_col = c
                break
        if header_row is not None:
            break

    def is_merged_cell(sheet, row, col):
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return True
        return False

    def get_merged_range(sheet, row, col):
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return merged_range
        return None

    def canonical_case_name(raw_name):
        if raw_name is None:
            return None
        name = str(raw_name).strip().casefold()
        if name == "reference":
            return "reference"

        name = re.sub(r"\s+", "_", name)
        if "_" not in name:
            return name

        prefix, alpha_token = name.rsplit("_", 1)
        token = alpha_token.replace(",", ".")
        if token == "uncapped":
            return f"{prefix}_uncapped"

        token = re.sub(r"(?<=\d)-(?!$)(?=\d)", ".", token)
        try:
            value = float(token)
        except ValueError:
            return name

        if value > 1.5:
            value = value / 100.0

        return f"{prefix}_{value:g}"

    variable_header_row = header_row

    if variable_header_row is None:
        variable_col = marker_col
        variable_start_row = marker_row + 1
    else:
        variable_start_row = variable_header_row + 1

    variables = []
    row = variable_start_row
    while True:
        val = ws.cell(row=row, column=variable_col).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            break
        variables.append((row, str(val)))
        row += 1

    title_merge = get_merged_range(ws, marker_row, marker_col)
    first_var_row = variables[0][0] if variables else variable_start_row
    var_merge = get_merged_range(ws, first_var_row, variable_col)

    if var_merge is not None:
        start_col = var_merge.max_col + 1
    else:
        start_col = variable_col + 1

    end_col = title_merge.max_col if title_merge is not None else ws.max_column
    available_cols = list(range(start_col, end_col + 1))
    if not available_cols:
        raise ValueError("No available columns in OPTIMIZATION RESULTS table.")

    case_header_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=marker_col).value
        if isinstance(val, str) and val.strip().casefold() == "case_name":
            case_header_row = r
            break

    if case_header_row is None and variable_header_row is not None:
        best_row = None
        best_count = 0
        for r in range(marker_row + 1, search_end + 1):
            if r < 1 or r > ws.max_row:
                continue
            count = 0
            for c in available_cols:
                val = ws.cell(row=r, column=c).value
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    continue
                count += 1
            if count > best_count:
                best_count = count
                best_row = r

        if best_row is not None and best_count > 0:
            case_header_row = best_row

    if case_header_row is not None:
        header_cols = []
        for c in available_cols:
            val = ws.cell(row=case_header_row, column=c).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            header_cols.append(c)
        if header_cols:
            available_cols = header_cols

    if len(cases_sorted) > len(available_cols):
        print(
            f"Warning: {len(cases_sorted)} cases but only {len(available_cols)} columns available. "
            "Extra cases will be ignored."
        )

    header_map = {}
    unassigned_cols = list(available_cols)
    if case_header_row is not None:
        for col in available_cols:
            cell_val = ws.cell(row=case_header_row, column=col).value
            if cell_val is None or (isinstance(cell_val, str) and cell_val.strip() == ""):
                continue
            key = canonical_case_name(cell_val)
            if key and key not in header_map:
                header_map[key] = col
                if col in unassigned_cols:
                    unassigned_cols.remove(col)

    case_to_col = {}
    missing_cases = []
    for case in cases_sorted:
        key = canonical_case_name(case)
        if key in header_map:
            case_to_col[case] = header_map[key]
            if header_map[key] in unassigned_cols:
                unassigned_cols.remove(header_map[key])
        elif case_header_row is None and unassigned_cols:
            case_to_col[case] = unassigned_cols.pop(0)
        else:
            missing_cases.append(case)

    if missing_cases:
        print(
            "Warning: the following cases are not present in the CASE_NAME header row and were skipped: "
            + ", ".join(missing_cases)
        )

    if case_header_row is not None and not is_merged_cell(ws, case_header_row, start_col):
        for case, col in case_to_col.items():
            if is_merged_cell(ws, case_header_row, col):
                continue
            existing = ws.cell(row=case_header_row, column=col).value
            if existing is None or (isinstance(existing, str) and existing.strip() == ""):
                ws.cell(row=case_header_row, column=col, value=case)

    for case, target_col in case_to_col.items():
        values = cases_data.get(case, {})
        for var_row, var_name in variables:
            if is_merged_cell(ws, var_row, target_col):
                continue
            val = values.get(var_name)
            if val is None and isinstance(var_name, str):
                val = values.get(var_name.strip())
            ws.cell(row=var_row, column=target_col, value=val if val is not None else float("nan"))

    wb.save(saving_path)
    print(f"Synthesis file created at: {saving_path}")
    return saving_path

def read_cases_synthesis_dict(synthesis_path=SYNTHESIS_FILE_PATH):
    """
    Read the synthesis Excel file and return a nested dictionary keyed by CASE_NAME.
    Each case dictionary contains keys "TES_TYPE", "DMS_SHARE" and all variables
    listed in the OPTIMIZATION RESULTS table.
    Returns : dict
        Mapping of case name -> dict of values.
    """
    if not os.path.exists(synthesis_path):
        raise FileNotFoundError(f"Synthesis file not found: {synthesis_path}")

    from openpyxl import load_workbook

    wb = load_workbook(synthesis_path, data_only=True)

    def find_case_header():
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip().casefold() == "case_name":
                        return ws, cell.row
        return None, None

    def find_optimization_marker():
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip().casefold() == "optimization results":
                        return ws, cell.row, cell.column
        return None, None, None

    def find_variable_header(ws, marker_row):
        keywords = {"variable", "variables", "parametre", "parametres", "paramètre", "paramètres"}
        search_end = marker_row + 10
        for r in range(marker_row + 1, min(search_end, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and val.strip().casefold() in keywords:
                    return r, c
        return None, None

    marker_ws, marker_row, marker_col = find_optimization_marker()
    if marker_ws is None:
        raise ValueError("'OPTIMIZATION RESULTS' section not found in synthesis file.")

    variable_header_row, variable_col = find_variable_header(marker_ws, marker_row)
    if variable_header_row is None:
        variable_col = marker_col
        variable_start_row = marker_row + 1
    else:
        variable_start_row = variable_header_row + 1

    def get_merged_range(sheet, row, col):
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return merged_range
        return None

    title_merge = get_merged_range(marker_ws, marker_row, marker_col)
    var_merge = get_merged_range(marker_ws, variable_start_row, variable_col)
    start_col = var_merge.max_col + 1 if var_merge is not None else variable_col + 1
    end_col = title_merge.max_col if title_merge is not None else marker_ws.max_column
    available_cols = list(range(start_col, end_col + 1))

    case_ws = marker_ws
    case_row = None
    search_end = min(marker_row + 15, case_ws.max_row)
    for r in range(marker_row, search_end + 1):
        for c in range(1, case_ws.max_column + 1):
            val = case_ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().casefold() == "case_name":
                case_row = r
                break
        if case_row is not None:
            break

    if case_row is None:
        case_ws, case_row = find_case_header()
        if case_ws is None:
            raise ValueError("CASE_NAME header row not found in synthesis file.")

    case_columns = []
    case_names = []
    for c in range(1, case_ws.max_column + 1):
        val = case_ws.cell(row=case_row, column=c).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            continue
        if isinstance(val, str) and val.strip().casefold() == "case_name":
            continue
        case_columns.append(c)
        case_names.append(str(val).strip())

    if not case_columns:
        raise ValueError("No case columns found under CASE_NAME header.")

    variables = []
    row = variable_start_row
    started = False
    while row <= marker_ws.max_row:
        val = marker_ws.cell(row=row, column=variable_col).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            if started:
                break
            row += 1
            continue
        started = True
        variables.append(str(val).strip())
        row += 1

    results = {case: {"TES_TYPE": None, "DMS_SHARE": None} for case in case_names}

    def normalize_label(value):
        if value is None:
            return ""
        return re.sub(r"[\s_\-]+", "", str(value).strip().casefold())

    var_rows = {}
    canonical_rows = {}
    started = False
    for r in range(variable_start_row, marker_ws.max_row + 1):
        val = marker_ws.cell(row=r, column=variable_col).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            if started:
                break
            continue
        started = True
        label = str(val).strip()
        var_rows[label] = r
        canonical_rows[normalize_label(label)] = r

    for var_name, var_row in var_rows.items():
        for case, col in zip(case_names, case_columns):
            results[case][var_name] = marker_ws.cell(row=var_row, column=col).value

    def find_label_row(label_key):
        for r in range(1, marker_ws.max_row + 1):
            for c in range(1, marker_ws.max_column + 1):
                val = marker_ws.cell(row=r, column=c).value
                if normalize_label(val) == label_key:
                    return r
        return None

    def merged_cell_value(sheet, row, col):
        cell = sheet.cell(row=row, column=col)
        if cell.value is not None:
            return cell.value
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return None

    tes_row = canonical_rows.get("testype") or find_label_row("testype")
    dms_row = canonical_rows.get("dmsshare") or find_label_row("dmsshare")
    if tes_row is not None:
        for case, col in zip(case_names, case_columns):
            results[case]["TES_TYPE"] = merged_cell_value(marker_ws, tes_row, col)
    if dms_row is not None:
        for case, col in zip(case_names, case_columns):
            results[case]["DMS_SHARE"] = merged_cell_value(marker_ws, dms_row, col)

    return results

def plot_cases_comparison_per_TES_type(variable, input_df, TES_types="all", saving_dir=None):
    """
    Plot a variable versus DMS_SHARE for selected TES types.
    ----------
    variable : str
        Variable name to plot (e.g., "v_TES", "C_total").
    input_df : dict
        Dictionary returned by read_cases_synthesis_dict().
    TES_types : list or str
        List of TES types to plot, or "all" for every type.
    saving_dir : str or None
        Directory to save the plot. If None, the plot is only displayed.
    """
    if not isinstance(input_df, dict):
        raise TypeError("input_df must be a dict returned by read_cases_synthesis_dict().")

    if isinstance(TES_types, str) and TES_types.strip().casefold() == "all":
        tes_filter = None
    else:
        tes_filter = set(TES_types)

    def filter_allows_reference():
        if tes_filter is None:
            return True
        for item in tes_filter:
            if item == 0 or str(item).strip() == "0":
                return True
        return False

    def filter_allows_dms_only():
        if tes_filter is None:
            return True
        for item in tes_filter:
            if item == 0 or str(item).strip() == "0":
                return True
        return False

    def to_float(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip().replace(",", "."))
        except ValueError:
            return None

    marker_cycle = ["o", "^", "s", "D", "v", "P", "X", "*"]
    color_cycle = ["steelblue", "darkgreen", "purple", "brown", "gray", "olive", "teal"]

    def apply_units(value, var_name):
        if value is None:
            return None, ""
        if not isinstance(var_name, str) or var_name == "":
            return value, ""
        prefix = var_name.strip()[:1].upper()
        if prefix == "C":
            return value / 1_000_000.0, "M€"
        if prefix == "V":
            return value, "m" + r"$^3$"
        if prefix == "Z":
            return value, "kW"
        if prefix == "W":
            return value / 1_000.0, "MWh"
        return value, ""

    y_unit = ""
    grouped = {}
    reference_points = []
    dms_only_points = []

    for case_name, values in input_df.items():
        if not isinstance(values, dict):
            continue

        tes_type = values.get("TES_TYPE")
        dms_share = to_float(values.get("DMS_SHARE"))
        raw_y = values.get(variable)
        y_val, y_unit = apply_units(raw_y, variable)

        if y_val is None or dms_share is None:
            continue

        case_key = str(case_name).strip().casefold()
        if case_key == "reference":
            if filter_allows_reference():
                reference_points.append((dms_share, y_val, case_name))
            continue
        if case_key == "dms_only":
            if filter_allows_dms_only():
                dms_only_points.append((dms_share, y_val, case_name))
            continue

        if tes_filter is not None and tes_type not in tes_filter:
            continue

        grouped.setdefault(tes_type, []).append((dms_share, y_val, case_name))

    fig, ax = plt.subplots(figsize=(7, 4.5))

    for idx, (tes_type, points) in enumerate(sorted(grouped.items(), key=lambda x: str(x[0]))):
        points_sorted = sorted(points, key=lambda x: x[0])
        x_vals = [p[0] for p in points_sorted]
        y_vals = [p[1] for p in points_sorted]
        marker = marker_cycle[idx % len(marker_cycle)]
        color = color_cycle[idx % len(color_cycle)]
        ax.scatter(x_vals, y_vals, marker=marker, color=color, s=55, label=f"TES_TYPE={tes_type}")
        if len(x_vals) > 1:
            ax.plot(x_vals, y_vals, linestyle="--", color=color, linewidth=1)

    if reference_points:
        ref_x = [p[0] for p in reference_points]
        ref_y = [p[1] for p in reference_points]
        ax.scatter(ref_x, ref_y, marker="s", color="red", s=70, label="reference")

    if dms_only_points:
        dms_x = [p[0] for p in dms_only_points]
        dms_y = [p[1] for p in dms_only_points]
        ax.scatter(dms_x, dms_y, marker="x", color="orange", s=70, label="DMS_only")

    ax.set_xlabel(r"$\alpha_{\max}$")
    ylabel = f"{variable} ({y_unit})" if y_unit else variable
    ax.set_ylabel(ylabel)
    ax.set_title(f"{variable} vs DMS_SHARE")
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=8, ncol=2)

    plt.tight_layout()

    if saving_dir is not None:
        os.makedirs(saving_dir, exist_ok=True)
        fig_path = os.path.join(saving_dir, f"cases_comparison_{variable}.png")
        plt.savefig(fig_path, dpi=300)
        print(f"Plot saved to: {fig_path}")

    plt.show()


def plot_cases_comparison_per_case_names(variable, input_df, case_names="all", labels=None, saving_dir=None):
    """
    Plot a variable versus DMS_SHARE for selected case names.
    ----------
    variable : str
        Variable name to plot (e.g., "v_TES", "C_total").
    input_df : dict
        Dictionary returned by read_cases_synthesis_dict().
    case_names : list or str
        List of case names to plot, or "all" for every case.
    labels : list or None
        Display names for each case, in the same order as case_names.
    saving_dir : str or None
        Directory to save the plot. If None, the plot is only displayed.
    """
    if not isinstance(input_df, dict):
        raise TypeError("input_df must be a dict returned by read_cases_synthesis_dict().")

    if isinstance(case_names, str) and case_names.strip().casefold() == "all":
        case_filter = None
        if labels is not None:
            raise ValueError("labels cannot be used when case_names is 'all'.")
        case_label_map = None
    else:
        normalized_names = [str(name).strip().casefold() for name in case_names]
        case_filter = set(normalized_names)
        if labels is not None:
            if len(labels) != len(normalized_names):
                raise ValueError("labels must have the same length as case_names.")
            case_label_map = {
                case_key: str(label) if label is not None else ""
                for case_key, label in zip(normalized_names, labels)
            }
        else:
            case_label_map = None

    def to_float(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip().replace(",", "."))
        except ValueError:
            return None

    marker_cycle = ["o", "^", "s", "D", "v", "P", "X", "*"]
    color_cycle = ["steelblue", "darkgreen", "purple", "brown", "gray", "olive", "teal"]

    def apply_units(value, var_name):
        if value is None:
            return None, ""
        if not isinstance(var_name, str) or var_name == "":
            return value, ""
        prefix = var_name.strip()[:1].upper()
        if prefix == "C":
            return value / 1_000_000.0, "M€"
        if prefix == "V":
            return value, "m" + r"$^3$"
        if prefix == "Z":
            return value, "kW"
        if prefix == "W":
            return value / 1_000.0, "MWh"
        return value, ""

    y_unit = ""
    grouped = {}
    reference_points = []
    dms_only_points = []

    for case_name, values in input_df.items():
        if not isinstance(values, dict):
            continue

        case_key = str(case_name).strip().casefold()
        if case_filter is not None and case_key not in case_filter:
            continue

        tes_type = values.get("TES_TYPE")
        dms_share = to_float(values.get("DMS_SHARE"))
        raw_y = values.get(variable)
        y_val, y_unit = apply_units(raw_y, variable)

        if y_val is None or dms_share is None:
            continue

        if case_key == "reference":
            display_label = case_label_map.get(case_key) if case_label_map is not None else "reference"
            reference_points.append((dms_share, y_val, display_label))
            continue
        if case_key == "dms_only":
            display_label = case_label_map.get(case_key) if case_label_map is not None else "DMS_only"
            dms_only_points.append((dms_share, y_val, display_label))
            continue

        display_label = case_label_map.get(case_key) if case_label_map is not None else case_name
        grouped.setdefault(tes_type, []).append((dms_share, y_val, case_name, display_label))

    fig, ax = plt.subplots(figsize=(7, 4.5))

    for idx, (tes_type, points) in enumerate(sorted(grouped.items(), key=lambda x: str(x[0]))):
        points_sorted = sorted(points, key=lambda x: x[0])
        x_vals = [p[0] for p in points_sorted]
        y_vals = [p[1] for p in points_sorted]
        marker = marker_cycle[idx % len(marker_cycle)]
        color = color_cycle[idx % len(color_cycle)]
        if case_label_map is None:
            ax.scatter(x_vals, y_vals, marker=marker, color=color, s=55, label=f"TES_TYPE={tes_type}")
        else:
            ax.scatter(x_vals, y_vals, marker=marker, color=color, s=55)
        if len(x_vals) > 1:
            ax.plot(x_vals, y_vals, linestyle="--", color=color, linewidth=1)

        if case_label_map is not None:
            for x_val, y_val, _case_name, display_label in points_sorted:
                ax.scatter([x_val], [y_val], marker=marker, color=color, s=55, label=display_label)

    if reference_points:
        ref_x = [p[0] for p in reference_points]
        ref_y = [p[1] for p in reference_points]
        ref_labels = [p[2] for p in reference_points]
        if case_label_map is None:
            ax.scatter(ref_x, ref_y, marker="s", color="red", s=70, label="reference")
        else:
            for x_val, y_val, label in zip(ref_x, ref_y, ref_labels):
                ax.scatter([x_val], [y_val], marker="s", color="red", s=70, label=label)

    if dms_only_points:
        dms_x = [p[0] for p in dms_only_points]
        dms_y = [p[1] for p in dms_only_points]
        dms_labels = [p[2] for p in dms_only_points]
        if case_label_map is None:
            ax.scatter(dms_x, dms_y, marker="x", color="orange", s=70, label="DMS_only")
        else:
            for x_val, y_val, label in zip(dms_x, dms_y, dms_labels):
                ax.scatter([x_val], [y_val], marker="x", color="orange", s=70, label=label)

    ax.set_xlabel(r"$\alpha_{\max}$")
    ylabel = f"{variable} ({y_unit})" if y_unit else variable
    ax.set_ylabel(ylabel)
    ax.set_title(f"{variable} vs DMS_SHARE")
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=8, ncol=2)

    plt.tight_layout()

    if saving_dir is not None:
        os.makedirs(saving_dir, exist_ok=True)
        fig_path = os.path.join(saving_dir, f"cases_comparison_{variable}.png")
        plt.savefig(fig_path, dpi=300)
        print(f"Plot saved to: {fig_path}")

    plt.show()


###################################### EXECUTION ######################################

compute_case_studies_synthesis()

dico = read_cases_synthesis_dict()
VARIABLES_TO_PLOT = ["C_total", "v_TES", "w_tot", "z_comp_DMS","z_comp_CO2"]

# TES_types_0 = [0, 'mono2']
# TES_types_0 = ['mono1', 'mono2', 'mono3', 'mono4', 'multi1', 'multi2', 'multi3']
case_names_0 = ['reference', 'DMS_only_uncapped', 'mono2_45']

# for var in VARIABLES_TO_PLOT:
    # plot_cases_comparison_per_TES_type(var, input_df=dico, TES_types=TES_types_0, saving_dir=PLOTS_SAVING_DIR)
    # plot_cases_comparison_per_case_names(var, input_df=dico, case_names=case_names_0, saving_dir=PLOTS_SAVING_DIR)

